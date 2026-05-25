from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BASELINE_DIR = PROJECT_ROOT / "outputs" / "forecast_benchmark" / "tables"
ENHANCED_DIR = PROJECT_ROOT / "outputs" / "enhanced_forecast_benchmark" / "tables"
BOOTSTRAP_DIR = PROJECT_ROOT / "outputs" / "main_benchmark_bootstrap" / "tables"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "publication_assets" / "tables"

OUTPUT_XLSX = OUTPUT_DIR / "table_compact_benchmark_stability.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "table_compact_benchmark_stability.csv"
SPLIT_OUTPUTS = {
    "Daily max AQI": {
        "xlsx": OUTPUT_DIR / "table_compact_benchmark_stability_aqi.xlsx",
        "csv": OUTPUT_DIR / "table_compact_benchmark_stability_aqi.csv",
        "root_copy": PROJECT_ROOT / "outputs" / "Benchmark stability AQI.xlsx",
    },
    "Daily mean PM2.5": {
        "xlsx": OUTPUT_DIR / "table_compact_benchmark_stability_pm25.xlsx",
        "csv": OUTPUT_DIR / "table_compact_benchmark_stability_pm25.csv",
        "root_copy": PROJECT_ROOT / "outputs" / "Benchmark stability PM2.5.xlsx",
    },
}
NARROW_SPLIT_OUTPUTS = {
    "Daily max AQI": {
        "xlsx": OUTPUT_DIR / "table_narrow_stability_aqi.xlsx",
        "csv": OUTPUT_DIR / "table_narrow_stability_aqi.csv",
        "root_copy": PROJECT_ROOT / "outputs" / "Benchmark stability AQI narrow.xlsx",
    },
    "Daily mean PM2.5": {
        "xlsx": OUTPUT_DIR / "table_narrow_stability_pm25.xlsx",
        "csv": OUTPUT_DIR / "table_narrow_stability_pm25.csv",
        "root_copy": PROJECT_ROOT / "outputs" / "Benchmark stability PM2.5 narrow.xlsx",
    },
}

TARGET_LABELS = {
    "european_aqi_max": "Daily max AQI",
    "pm2_5_mean": "Daily mean PM2.5",
}
MODEL_SHORT = {
    "Naive": "Naive",
    "Seasonal Naive": "SNaive-7",
    "Drift": "Drift",
    "Holt Damped": "Holt",
    "ARIMA": "ARIMA",
    "Random Forest Lag7": "RF-Lag7",
    "HistGBM Global Features": "HistGBM",
    "Ridge Global Features": "Ridge",
    "Persistence Current": "Persist",
}


def build_compact_table() -> pd.DataFrame:
    comparison = pd.read_csv(ENHANCED_DIR / "comparison_vs_baseline.csv")
    baseline_metrics = pd.read_csv(BASELINE_DIR / "metrics_by_city.csv")
    enhanced_metrics = pd.read_csv(ENHANCED_DIR / "metrics_by_city.csv")
    bootstrap = pd.read_csv(BOOTSTRAP_DIR / "bootstrap_city_level_summary.csv")

    rows: list[dict[str, object]] = []
    for record in comparison.itertuples(index=False):
        baseline_subset = baseline_metrics[
            (baseline_metrics["target"] == record.target)
            & (baseline_metrics["horizon_days"] == record.horizon_days)
            & (baseline_metrics["model"] == record.baseline_best_model)
        ][["city", "mae"]].rename(columns={"mae": "baseline_mae"})

        enhanced_subset = enhanced_metrics[
            (enhanced_metrics["target"] == record.target)
            & (enhanced_metrics["horizon_days"] == record.horizon_days)
            & (enhanced_metrics["model"] == record.enhanced_best_model)
        ][["city", "mae"]].rename(columns={"mae": "enhanced_mae"})

        merged = enhanced_subset.merge(baseline_subset, on="city", how="inner")
        enhanced_city_wins = int((merged["enhanced_mae"] < merged["baseline_mae"]).sum())
        baseline_city_wins = int((merged["enhanced_mae"] > merged["baseline_mae"]).sum())

        ci_row = bootstrap[
            (bootstrap["target"] == record.target)
            & (bootstrap["horizon_days"] == record.horizon_days)
        ].iloc[0]

        rows.append(
            {
                "Target": TARGET_LABELS[record.target],
                "H": int(record.horizon_days),
                "Enhanced": MODEL_SHORT[record.enhanced_best_model],
                "Baseline": MODEL_SHORT[record.baseline_best_model],
                "MAE (E)": round(float(record.enhanced_mean_mae), 4),
                "MAE (B)": round(float(record.baseline_mean_mae), 4),
                "Delta B-E": round(float(record.mae_improvement), 4),
                "95% CI (B-E)": (
                    f"[{float(ci_row['ci_lower_95_baseline_minus_enhanced_mae']):.4f}, "
                    f"{float(ci_row['ci_upper_95_baseline_minus_enhanced_mae']):.4f}]"
                ),
                "City wins E:B": f"{enhanced_city_wins}:{baseline_city_wins}",
            }
        )

    return pd.DataFrame(rows)


def write_workbook(table: pd.DataFrame, output_path: Path, title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stability"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(table.columns))

    headers = list(table.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(table.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [2, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    note_row = len(table) + 4
    ws.cell(
        row=note_row,
        column=1,
        value="Note: City wins summarize which model had lower MAE in each of the 8 city-level comparisons.",
    )
    ws.cell(
        row=note_row + 1,
        column=1,
        value="The confidence interval is a paired city-level 95% bootstrap interval for baseline minus enhanced MAE.",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=len(headers))
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.cell(row=note_row + 1, column=1).alignment = Alignment(wrap_text=True)

    widths = {
        "A": 19,
        "B": 5,
        "C": 12,
        "D": 12,
        "E": 10,
        "F": 10,
        "G": 11,
        "H": 20,
        "I": 12,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows(min_row=2, max_row=2 + len(table), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "$1:$2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_narrow_table(table: pd.DataFrame) -> pd.DataFrame:
    narrow = table.copy()
    narrow["Comparator"] = narrow["Baseline"]
    narrow["95% CI"] = narrow["95% CI (B-E)"]
    narrow["Wins E:B"] = narrow["City wins E:B"]
    return narrow[["H", "Comparator", "Delta B-E", "95% CI", "Wins E:B"]]


def write_narrow_workbook(table: pd.DataFrame, output_path: Path, title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stability"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(table.columns))

    for col_idx, header in enumerate(table.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(table.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    note_row = len(table) + 4
    ws.cell(
        row=note_row,
        column=1,
        value="Delta B-E = baseline minus enhanced MAE; positive values favor the enhanced benchmark.",
    )
    ws.cell(
        row=note_row + 1,
        column=1,
        value="Wins E:B summarize city-level wins; mean MAEs are already reported in the main benchmark table.",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(table.columns))
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=len(table.columns))
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.cell(row=note_row + 1, column=1).alignment = Alignment(wrap_text=True)

    widths = {
        "A": 6,
        "B": 12,
        "C": 11,
        "D": 20,
        "E": 10,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows(min_row=2, max_row=2 + len(table), min_col=1, max_col=len(table.columns)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "$1:$2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    table = build_compact_table()
    table.to_csv(OUTPUT_CSV, index=False)
    write_workbook(
        table,
        OUTPUT_XLSX,
        "Benchmark stability summary (positive Delta B-E favors enhanced benchmark)",
    )

    saved_paths = [OUTPUT_XLSX, OUTPUT_CSV]
    for target_name, paths in SPLIT_OUTPUTS.items():
        target_table = table[table["Target"] == target_name].reset_index(drop=True)
        target_table.to_csv(paths["csv"], index=False)
        write_workbook(
            target_table,
            paths["xlsx"],
            f"{target_name} stability summary (positive Delta B-E favors enhanced benchmark)",
        )
        paths["root_copy"].write_bytes(paths["xlsx"].read_bytes())
        saved_paths.extend([paths["xlsx"], paths["csv"], paths["root_copy"]])

    for target_name, paths in NARROW_SPLIT_OUTPUTS.items():
        target_table = build_narrow_table(table[table["Target"] == target_name].reset_index(drop=True))
        target_table.to_csv(paths["csv"], index=False)
        write_narrow_workbook(
            target_table,
            paths["xlsx"],
            f"{target_name} stability evidence (HistGBM vs best local baseline)",
        )
        paths["root_copy"].write_bytes(paths["xlsx"].read_bytes())
        saved_paths.extend([paths["xlsx"], paths["csv"], paths["root_copy"]])

    print("Saved compact stability table to:")
    for path in saved_paths:
        print(f"  - {path}")


if __name__ == "__main__":
    main()
